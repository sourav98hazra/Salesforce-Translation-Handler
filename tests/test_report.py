"""Tests for the validation report export module."""

from __future__ import annotations

import csv
import json
from pathlib import Path

from stx.model import Document, Entry
from stx.report import export_csv, export_html, export_json
from stx.validate import ValidationIssue, ValidationReport, validate_document


def _make_report_with_issues() -> ValidationReport:
    """Create a report with known issues for testing."""
    doc = Document(
        entries=[
            # Duplicate keys -> error
            Entry(key="CustomLabel.Greeting", label="Hello", translation="Bonjour"),
            Entry(key="CustomLabel.Greeting", label="Hi", translation="Salut"),
            # Length limit violation -> error
            Entry(key="CustomField.A.Foo.FieldLabel", label="hi", translation="x" * 200),
            # Token drift -> error
            Entry(
                key="CustomLabel.Msg",
                label="Hello {!User.Name}",
                translation="Bonjour",
            ),
            # HTML mismatch -> warning
            Entry(
                key="ButtonOrLink.Body",
                label="<p>Hello</p>",
                translation="Bonjour",
            ),
        ],
    )
    return validate_document(doc)


def _make_empty_report() -> ValidationReport:
    """Create a report with zero issues."""
    doc = Document(
        entries=[
            Entry(key="CustomLabel.A", label="Hello", translation="Bonjour"),
        ]
    )
    return validate_document(doc)


# --- CSV tests ---


def test_csv_export_has_correct_headers(tmp_path: Path) -> None:
    report = _make_report_with_issues()
    out = tmp_path / "report.csv"
    export_csv(report, out)
    assert out.exists()

    with out.open(encoding="utf-8") as f:
        reader = csv.reader(f)
        rows = list(reader)

    # First row is summary comment
    assert rows[0][0].startswith("# Summary:")
    # Second row is header
    assert rows[1] == ["severity", "category", "component", "key", "message"]


def test_csv_export_row_count(tmp_path: Path) -> None:
    report = _make_report_with_issues()
    out = tmp_path / "report.csv"
    export_csv(report, out)

    with out.open(encoding="utf-8") as f:
        reader = csv.reader(f)
        rows = list(reader)

    # summary + header + one per issue
    assert len(rows) == 2 + len(report.issues)


def test_csv_export_empty_report(tmp_path: Path) -> None:
    report = _make_empty_report()
    out = tmp_path / "report.csv"
    export_csv(report, out)
    assert out.exists()

    with out.open(encoding="utf-8") as f:
        reader = csv.reader(f)
        rows = list(reader)

    # summary + header, no data rows
    assert len(rows) == 2
    assert "0 error(s)" in rows[0][0]


# --- JSON tests ---


def test_json_export_valid_json(tmp_path: Path) -> None:
    report = _make_report_with_issues()
    out = tmp_path / "report.json"
    export_json(report, out)
    assert out.exists()

    data = json.loads(out.read_text(encoding="utf-8"))
    assert "summary" in data
    assert "issues" in data
    assert "issues_by_category" in data


def test_json_export_summary_counts(tmp_path: Path) -> None:
    report = _make_report_with_issues()
    out = tmp_path / "report.json"
    export_json(report, out)

    data = json.loads(out.read_text(encoding="utf-8"))
    assert data["summary"]["errors"] == len(report.errors)
    assert data["summary"]["warnings"] == len(report.warnings)
    assert data["summary"]["total"] == len(report.issues)


def test_json_export_issues_grouped_by_category(tmp_path: Path) -> None:
    report = _make_report_with_issues()
    out = tmp_path / "report.json"
    export_json(report, out)

    data = json.loads(out.read_text(encoding="utf-8"))
    by_cat = data["issues_by_category"]
    # We know the report has at least duplicate_key and length_limit categories
    assert "duplicate_key" in by_cat
    assert "length_limit" in by_cat


def test_json_export_empty_report(tmp_path: Path) -> None:
    report = _make_empty_report()
    out = tmp_path / "report.json"
    export_json(report, out)
    assert out.exists()

    data = json.loads(out.read_text(encoding="utf-8"))
    assert data["summary"]["errors"] == 0
    assert data["summary"]["warnings"] == 0
    assert data["summary"]["total"] == 0
    assert data["issues"] == []
    assert data["issues_by_category"] == {}


# --- HTML tests ---


def test_html_export_contains_table(tmp_path: Path) -> None:
    report = _make_report_with_issues()
    out = tmp_path / "report.html"
    export_html(report, out)
    assert out.exists()

    content = out.read_text(encoding="utf-8")
    assert "<table>" in content
    assert "<th>Severity</th>" in content
    assert "<th>Component</th>" in content
    assert "duplicate_key" in content


def test_html_export_has_summary(tmp_path: Path) -> None:
    report = _make_report_with_issues()
    out = tmp_path / "report.html"
    export_html(report, out)

    content = out.read_text(encoding="utf-8")
    assert f"Errors: {len(report.errors)}" in content
    assert f"Warnings: {len(report.warnings)}" in content


def test_html_export_empty_report(tmp_path: Path) -> None:
    report = _make_empty_report()
    out = tmp_path / "report.html"
    export_html(report, out)
    assert out.exists()

    content = out.read_text(encoding="utf-8")
    assert "Errors: 0" in content
    assert "Warnings: 0" in content
    assert "<table>" in content


# --- Component field tests ---


def test_component_field_populated_in_validation_issue() -> None:
    doc = Document(
        entries=[
            Entry(
                key="CustomField.Account.Name.FieldLabel",
                label="Name",
                translation="x" * 200,
            ),
        ]
    )
    report = validate_document(doc)
    assert report.issues
    for issue in report.issues:
        assert issue.component == "CustomField"


def test_component_field_populated_for_duplicate_keys() -> None:
    doc = Document(
        entries=[
            Entry(key="ButtonOrLink.X", label="Click"),
            Entry(key="ButtonOrLink.X", label="Press"),
        ]
    )
    report = validate_document(doc)
    assert report.issues
    assert report.issues[0].component == "ButtonOrLink"


def test_component_field_in_as_dict() -> None:
    issue = ValidationIssue(
        category="test",
        severity="error",
        key="CustomLabel.X",
        message="test message",
        component="CustomLabel",
    )
    d = issue.as_dict()
    assert d["component"] == "CustomLabel"


def test_component_field_defaults_to_empty_string() -> None:
    """Backward compatibility: component defaults to empty string."""
    issue = ValidationIssue(
        category="test",
        severity="error",
        key="CustomLabel.X",
        message="test message",
    )
    assert issue.component == ""


def test_report_grouping_matches_categories(tmp_path: Path) -> None:
    """JSON export groups match the report's by_category method."""
    report = _make_report_with_issues()
    out = tmp_path / "report.json"
    export_json(report, out)

    data = json.loads(out.read_text(encoding="utf-8"))
    report_categories = report.by_category()
    for cat, issues in report_categories.items():
        assert cat in data["issues_by_category"]
        assert len(data["issues_by_category"][cat]) == len(issues)


# --- XLSX tests ---


def test_export_xlsx(tmp_path):
    """export_xlsx creates a workbook with Summary + per-category sheets."""
    from stx.report import export_xlsx

    report = _make_report_with_issues()
    out = tmp_path / "report.xlsx"
    export_xlsx(report, out)
    assert out.exists()
    from openpyxl import load_workbook

    wb = load_workbook(out)
    # Summary sheet is always present
    assert "Summary" in wb.sheetnames
    # Per-category sheets exist for each issue category
    categories = report.by_category()
    for cat in categories:
        assert cat[:31] in wb.sheetnames
    # Check Summary sheet has header
    ws = wb["Summary"]
    assert ws.cell(1, 1).value == "Validation Report"
    # Check that a category sheet has the right columns
    first_cat = sorted(categories.keys())[0]
    ws_cat = wb[first_cat[:31]]
    assert ws_cat.cell(1, 1).value == "#"
    assert ws_cat.cell(1, 2).value == "Severity"
    assert ws_cat.cell(1, 7).value == "Message"
    assert ws_cat.max_row > 1  # at least one data row


def test_export_xlsx_with_fixes(tmp_path):
    """export_xlsx includes fix data when fixes_applied is provided."""
    from stx.report import export_xlsx

    report = _make_report_with_issues()
    fixes = [
        {
            "key": "CustomLabel.Msg",
            "label": "Hello {!User.Name}",
            "previous_translation": "Bonjour",
            "fixed_translation": "Bonjour {!User.Name}",
            "issue_category": "token_drift",
            "fix_description": "Restored missing placeholder {!User.Name}",
        }
    ]
    out = tmp_path / "report.xlsx"
    export_xlsx(report, out, fixes_applied=fixes)
    from openpyxl import load_workbook

    wb = load_workbook(out)
    assert "Fixes Applied" in wb.sheetnames
    ws = wb["Fixes Applied"]
    assert ws.cell(1, 1).value == "Key"
    assert ws.max_row == 2  # header + 1 fix row
    assert ws.cell(2, 1).value == "CustomLabel.Msg"


def test_export_xlsx_summary_with_document_stats(tmp_path):
    """export_xlsx Summary sheet includes document stats when provided."""
    from stx.report import export_xlsx

    report = _make_report_with_issues()
    stats = {"total": 500, "translated": 450, "untranslated": 50}
    out = tmp_path / "report.xlsx"
    export_xlsx(report, out, document_stats=stats, document_name="test_file.xlsx")
    from openpyxl import load_workbook

    wb = load_workbook(out)
    ws = wb["Summary"]
    # Find document name and stats in the sheet
    values = [ws.cell(r, 1).value for r in range(1, ws.max_row + 1)]
    assert "Document" in values
    assert "Total Rows" in values
    # Check the document name value
    doc_row = values.index("Document") + 1
    assert ws.cell(doc_row, 2).value == "test_file.xlsx"


def test_export_xlsx_category_breakdown(tmp_path):
    """export_xlsx Summary sheet has category breakdown with totals."""
    from stx.report import export_xlsx

    report = _make_report_with_issues()
    out = tmp_path / "report.xlsx"
    export_xlsx(report, out)
    from openpyxl import load_workbook

    wb = load_workbook(out)
    ws = wb["Summary"]
    # Find the Category header row
    cat_row = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 1).value == "Category":
            cat_row = r
            break
    assert cat_row is not None
    # The row after the last category should be TOTAL
    # Find TOTAL row
    total_row = None
    for r in range(cat_row + 1, ws.max_row + 1):
        if ws.cell(r, 1).value == "TOTAL":
            total_row = r
            break
    assert total_row is not None
    # TOTAL errors + warnings should match report totals
    assert ws.cell(total_row, 2).value == len(report.errors)
    assert ws.cell(total_row, 3).value == len(report.warnings)
