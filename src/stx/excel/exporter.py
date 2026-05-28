"""Export an STF :class:`Document` to a structured ``.xlsx`` workbook.

The output matches what ``stftoexcel_v2.ps1`` produces:

* One sheet per ``ComponentType_Status`` group, with columns ``Key``,
  ``Label``, ``Translation``.
* Sheet names are truncated to 28 characters and disambiguated with a
  ``_1``, ``_2`` ... numeric suffix on collision (Excel's hard limit is
  31 chars; the suffix budget keeps us within that).
* A ``Content Details`` index sheet listing every component sheet.

After translation, callers may invoke :func:`write_translation_audit_sheets`
to append the ``Translation_Summary`` and ``Translation_Status_Log``
sheets emitted by the legacy Python translator.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable, List, Mapping

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ..model import Document, Entry

# Excel sheet name hard limit (Excel itself rejects > 31 chars).
_MAX_SHEET_NAME = 31
# Truncation budget that leaves room for a "_NN" disambiguation suffix.
_BASE_SHEET_NAME_BUDGET = 28

ENTRY_COLUMNS = ["Key", "Label", "Translation"]
CONTENT_DETAILS_COLUMNS = [
    "SheetName",
    "SavedAs",
    "ComponentType",
    "TranslationStatus",
    "TotalRecords",
]
CONTENT_DETAILS_SHEET = "Content Details"

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")


@dataclass
class ExcelExportResult:
    """The output of :func:`export_document_to_excel`.

    Attributes
    ----------
    path:
        File system location of the workbook.
    sheet_name_map:
        Mapping from logical ``ComponentType_Status`` name to the actual
        sheet name written into the workbook (post-truncation).
    sheets_written:
        Ordered list of every sheet name that was written.
    """

    path: Path
    sheet_name_map: dict = field(default_factory=dict)
    sheets_written: List[str] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def export_document_to_excel(doc: Document, output_path: Path | str) -> ExcelExportResult:
    """Write ``doc`` to ``output_path`` as an organized workbook."""

    target = Path(output_path)
    target.parent.mkdir(parents=True, exist_ok=True)

    grouped = _group_by_logical_sheet(doc.entries)

    wb = Workbook()
    # Workbook() creates a single default sheet which we'll repurpose for the
    # first group, or remove if we end up with no groups.
    default_ws = wb.active
    default_ws.title = "_tmp_"

    sheet_name_map: dict[str, str] = {}
    used_names: set[str] = set()
    sheets_written: list[str] = []

    for logical_name, entries in grouped.items():
        actual_name = _allocate_sheet_name(logical_name, used_names)
        used_names.add(actual_name)
        sheet_name_map[logical_name] = actual_name

        if default_ws and default_ws.title == "_tmp_":
            ws = default_ws
            ws.title = actual_name
            default_ws = None
        else:
            ws = wb.create_sheet(actual_name)

        _write_entry_sheet(ws, entries)
        sheets_written.append(actual_name)

    # Build the Content Details index sheet last so it's appended at the end
    # (matching stftoexcel_v2.ps1 ordering).
    if default_ws and default_ws.title == "_tmp_":
        # Empty document -- repurpose the default sheet for Content Details.
        details_ws = default_ws
        details_ws.title = CONTENT_DETAILS_SHEET
    else:
        details_ws = wb.create_sheet(CONTENT_DETAILS_SHEET)

    _write_content_details(details_ws, grouped, sheet_name_map)
    sheets_written.append(CONTENT_DETAILS_SHEET)

    wb.save(target)

    return ExcelExportResult(
        path=target,
        sheet_name_map=sheet_name_map,
        sheets_written=sheets_written,
    )


def write_translation_audit_sheets(
    workbook_path: Path | str,
    summary_rows: Iterable[Mapping[str, object]],
    status_rows: Iterable[Mapping[str, object]],
) -> None:
    """Append (or replace) the audit sheets to an existing workbook.

    This is invoked by the translation phase to reproduce the
    ``Translation_Summary`` and ``Translation_Status_Log`` sheets emitted
    by the legacy ``translate_excel_fixed.py`` script.
    """

    from openpyxl import load_workbook  # local import: avoids slow module load

    target = Path(workbook_path)
    wb = load_workbook(target)

    summary_columns = ["Sheet Name", "Total Rows", "Translated Rows", "Skipped Rows"]
    status_columns = ["Sheet Name", "Row Index", "Key", "Label", "Status"]

    _replace_sheet_with_dict_rows(wb, "Translation_Summary", summary_columns, summary_rows)
    _replace_sheet_with_dict_rows(wb, "Translation_Status_Log", status_columns, status_rows)

    wb.save(target)


# ---------------------------------------------------------------------------
# Internals
# ---------------------------------------------------------------------------

def _group_by_logical_sheet(entries: Iterable[Entry]) -> "dict[str, list[Entry]]":
    """Group entries by their ``logical_sheet_name`` while preserving order."""

    grouped: dict[str, list[Entry]] = {}
    for entry in entries:
        grouped.setdefault(entry.logical_sheet_name, []).append(entry)
    return grouped


def _allocate_sheet_name(logical_name: str, used: set[str]) -> str:
    """Mirror the truncation/collision logic from ``stftoexcel_v2.ps1``.

    Also strips characters Excel forbids in sheet names (``: \\ / ? * [ ]``)
    and refuses leading / trailing single quotes.  The legacy PowerShell
    script relied on Salesforce metadata never producing those characters,
    but defending against them makes the import path safer for hand-edited
    workbooks too.
    """

    sanitized = _sanitize_sheet_name(logical_name)
    base = sanitized[:_BASE_SHEET_NAME_BUDGET] or "Sheet"
    candidate = base
    counter = 1
    while candidate in used:
        candidate = f"{base}_{counter}"
        # Safety check: never exceed Excel's 31-char ceiling.
        if len(candidate) > _MAX_SHEET_NAME:
            base = base[: _MAX_SHEET_NAME - len(f"_{counter}")]
            candidate = f"{base}_{counter}"
        counter += 1
    return candidate


# Excel forbids these characters in sheet names.
_FORBIDDEN_SHEET_CHARS = re.compile(r"[:\\/?*\[\]]")


def _sanitize_sheet_name(name: str) -> str:
    cleaned = _FORBIDDEN_SHEET_CHARS.sub("_", name)
    cleaned = cleaned.strip().strip("'")
    return cleaned or "Sheet"


def _write_entry_sheet(ws: Worksheet, entries: List[Entry]) -> None:
    ws.append(ENTRY_COLUMNS)
    for entry in entries:
        ws.append([
            _safe_excel_text(entry.key),
            _safe_excel_text(entry.label),
            _safe_excel_text(entry.translation),
        ])
    _force_string_columns(ws, len(ENTRY_COLUMNS))
    _style_header(ws, len(ENTRY_COLUMNS))
    _autosize(ws, len(ENTRY_COLUMNS))


def _write_content_details(
    ws: Worksheet,
    grouped: Mapping[str, List[Entry]],
    sheet_name_map: Mapping[str, str],
) -> None:
    ws.append(CONTENT_DETAILS_COLUMNS)
    for logical_name, entries in grouped.items():
        component, _, status = logical_name.partition("_")
        ws.append([
            logical_name,
            sheet_name_map.get(logical_name, logical_name),
            component or "Unknown",
            status or "",
            len(entries),
        ])
    _style_header(ws, len(CONTENT_DETAILS_COLUMNS))
    _autosize(ws, len(CONTENT_DETAILS_COLUMNS))


def _replace_sheet_with_dict_rows(
    wb: Workbook,
    sheet_name: str,
    columns: List[str],
    rows: Iterable[Mapping[str, object]],
) -> None:
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ws.append(columns)
    for row in rows:
        ws.append([row.get(col, "") for col in columns])
    _style_header(ws, len(columns))
    _autosize(ws, len(columns))


def _style_header(ws: Worksheet, col_count: int) -> None:
    for col_idx in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
    ws.freeze_panes = "A2"


def _safe_excel_text(value: object) -> str:
    """Return ``value`` as a string safe to write into a worksheet cell.

    Mitigates two risks:

    1. **Formula injection** -- a label that begins with ``=``, ``+``,
       ``-`` or ``@`` is interpreted by Excel as a formula and may
       execute external commands when the workbook is opened.  We prefix
       such values with a single quote (Excel's standard convention for
       "this is plain text").
    2. **Type coercion** -- by always passing strings (rather than
       ``int``, ``float`` or ``datetime``) we prevent Excel from re-typing
       Salesforce keys like ``001`` or ``10:30`` as numbers / times.
    """
    if value is None:
        return ""
    text = str(value)
    if text and text[0] in {"=", "+", "-", "@"}:
        return "'" + text
    return text


def _force_string_columns(ws: Worksheet, col_count: int) -> None:
    """Set ``@`` (text) number-format on every data column.

    Belt-and-braces companion to :func:`_safe_excel_text`: even if a row
    is appended elsewhere, Excel won't re-interpret it as numeric.
    """
    from openpyxl.utils import get_column_letter

    for c in range(1, col_count + 1):
        ws.column_dimensions[get_column_letter(c)].number_format = "@"


def _autosize(ws: Worksheet, col_count: int, max_width: int = 80) -> None:
    """Compute a reasonable column width.

    openpyxl has no native autosize; we approximate by walking the rows
    and capping at ``max_width`` so a long label doesn't blow out the
    layout.
    """
    widths = [len(str(ws.cell(row=1, column=c).value or "")) for c in range(1, col_count + 1)]
    for row in ws.iter_rows(min_row=2, max_col=col_count, values_only=True):
        for idx, value in enumerate(row):
            if value is None:
                continue
            length = len(str(value))
            if length > widths[idx]:
                widths[idx] = length
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = min(width + 2, max_width)
