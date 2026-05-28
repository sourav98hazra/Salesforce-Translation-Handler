"""Read an organised translation workbook back into an STF :class:`Document`.

The importer is the inverse of :mod:`stx.excel.exporter`.  It uses the
``Content Details`` index sheet (when present) to determine which
component sheets to read, falling back to a permissive scan of every
sheet that has the expected ``Key`` / ``Label`` / ``Translation``
columns.

This dual mode means the importer can also consume manually edited
files (a common scenario when clients review translations in Excel and
hand the sheet back).
"""

from __future__ import annotations

from pathlib import Path
from typing import Iterable, List

from openpyxl import load_workbook

from ..model import Document, Entry
from .exporter import (
    CONTENT_DETAILS_COLUMNS,
    CONTENT_DETAILS_SHEET,
    ENTRY_COLUMNS,
)

_AUDIT_SHEETS = {"Translation_Summary", "Translation_Status_Log"}


def import_document_from_excel(
    path: Path | str,
    *,
    language: str | None = None,
    language_code: str | None = None,
    translation_type: str = "Metadata",
) -> Document:
    """Load an Excel workbook produced by the exporter (or hand-edited)."""

    wb = load_workbook(Path(path), data_only=True, read_only=True)

    sheet_names = _ordered_data_sheets(wb)
    entries: List[Entry] = []

    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        entries.extend(_read_entries_from_sheet(ws))

    return Document(
        language=language or "",
        language_code=language_code or "",
        stf_type="Bilingual",
        translation_type=translation_type,
        entries=entries,
    )


# ---------------------------------------------------------------------------
# Internals
# ---------------------------------------------------------------------------

def _ordered_data_sheets(wb) -> List[str]:
    """Return the list of sheets to read, in the order they should be read.

    Prefer the ``Content Details`` index when available so that we
    process sheets in the same order they were written.
    """

    if CONTENT_DETAILS_SHEET in wb.sheetnames:
        ws = wb[CONTENT_DETAILS_SHEET]
        rows = ws.iter_rows(values_only=True)
        try:
            header = list(next(rows))
        except StopIteration:
            header = []
        col_index = {name: idx for idx, name in enumerate(header) if name}

        saved_as_idx = col_index.get("SavedAs")
        if saved_as_idx is None:
            saved_as_idx = col_index.get("SheetName")

        if saved_as_idx is not None:
            ordered: list[str] = []
            for row in rows:
                if not row or saved_as_idx >= len(row):
                    continue
                name = row[saved_as_idx]
                if name and name in wb.sheetnames and name not in _AUDIT_SHEETS:
                    ordered.append(str(name))
            if ordered:
                return ordered

    # Fallback: every sheet that isn't an index/audit sheet.
    return [
        name
        for name in wb.sheetnames
        if name != CONTENT_DETAILS_SHEET and name not in _AUDIT_SHEETS
    ]


def _read_entries_from_sheet(ws) -> Iterable[Entry]:
    rows = ws.iter_rows(values_only=True)
    try:
        header_row = list(next(rows))
    except StopIteration:
        return []

    headers = [str(h).strip() if h is not None else "" for h in header_row]
    if not _has_entry_columns(headers):
        return []

    key_idx = headers.index("Key")
    label_idx = headers.index("Label")
    trans_idx = headers.index("Translation") if "Translation" in headers else None

    entries: list[Entry] = []
    for row in rows:
        if row is None:
            continue
        # Pad row to expected width to avoid IndexError on short/empty trailing cells.
        cells = list(row) + [None] * (max(key_idx, label_idx, trans_idx or 0) + 1 - len(row))
        key = _stringify(cells[key_idx])
        label = _stringify(cells[label_idx])
        translation = _stringify(cells[trans_idx]) if trans_idx is not None else ""
        if not key and not label:
            continue
        entries.append(Entry(key=key, label=label, translation=translation))
    return entries


def _has_entry_columns(headers: list[str]) -> bool:
    return all(col in headers for col in ENTRY_COLUMNS[:2])  # Key + Label required


def _stringify(value) -> str:
    if value is None:
        return ""
    text = str(value)
    # openpyxl returns NaN-ish placeholders sometimes from empty merged cells.
    if text.lower() in {"nan", "none"}:
        return ""
    # If the cell was previously protected with a leading apostrophe (our own
    # formula-injection guard), strip that single character back off so the
    # original value round-trips exactly.
    if text.startswith("'") and len(text) > 1 and text[1] in {"=", "+", "-", "@"}:
        return text[1:]
    return text


# Re-export for documentation discoverability.
__all__ = ["import_document_from_excel", "CONTENT_DETAILS_COLUMNS"]
