"""Export validation reports to CSV, JSON, HTML, and XLSX formats.

Each function takes a :class:`~stx.validate.ValidationReport` and writes
a formatted report to the given path.  An optional ``fixes_applied`` list
can be provided to include a "Fixes Applied" section showing what auto-fix
changed (before/after details).
"""

from __future__ import annotations

import csv
import json
from html import escape as html_escape
from pathlib import Path
from typing import List, Optional

from .validate import ValidationIssue, ValidationReport


def _component_from_key(key: str) -> str:
    """Extract component type from the first dot-segment of a key."""
    if "." in key:
        head = key.split(".", 1)[0]
        return head if head else "Unknown"
    return key or "Unknown"


def export_csv(
    report: ValidationReport,
    path: Path,
    fixes_applied: Optional[List[dict]] = None,
) -> None:
    """Write a CSV report with a summary row and one row per issue.

    If *fixes_applied* is provided, a second section (separated by a blank
    row) is appended with columns:
    Key, Label, Previous Translation, Fixed Translation, Issue Category, Fix Applied.
    """
    path = Path(path)
    error_count = len(report.errors)
    warning_count = len(report.warnings)

    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        # Summary row at top
        writer.writerow([
            f"# Summary: {error_count} error(s), {warning_count} warning(s), "
            f"{len(report.issues)} total issue(s)",
        ])
        # Header
        writer.writerow(["severity", "category", "component", "key", "message"])
        # Data rows
        for issue in report.issues:
            component = issue.component if issue.component else _component_from_key(issue.key)
            writer.writerow([
                issue.severity,
                issue.category,
                component,
                issue.key,
                issue.message,
            ])

        # Fixes Applied section
        if fixes_applied:
            writer.writerow([])  # blank separator row
            writer.writerow([
                f"# Fixes Applied: {len(fixes_applied)} fix(es)",
            ])
            writer.writerow([
                "Key", "Label", "Previous Translation",
                "Fixed Translation", "Issue Category", "Fix Applied",
            ])
            for fix in fixes_applied:
                writer.writerow([
                    fix.get("key", ""),
                    fix.get("label", ""),
                    fix.get("previous_translation", ""),
                    fix.get("fixed_translation", ""),
                    fix.get("issue_category", ""),
                    fix.get("fix_description", ""),
                ])


def export_json(
    report: ValidationReport,
    path: Path,
    fixes_applied: Optional[List[dict]] = None,
) -> None:
    """Write a JSON report with summary, issues grouped by category, and flat list.

    If *fixes_applied* is provided, an additional ``"fixes_applied"`` array
    is included in the output.
    """
    path = Path(path)
    error_count = len(report.errors)
    warning_count = len(report.warnings)

    issues_by_category: dict[str, list[dict]] = {}
    issues_flat: list[dict] = []

    for issue in report.issues:
        component = issue.component if issue.component else _component_from_key(issue.key)
        issue_dict = {
            "severity": issue.severity,
            "category": issue.category,
            "component": component,
            "key": issue.key,
            "message": issue.message,
        }
        issues_flat.append(issue_dict)
        issues_by_category.setdefault(issue.category, []).append(issue_dict)

    data: dict = {
        "summary": {
            "errors": error_count,
            "warnings": warning_count,
            "total": len(report.issues),
            "fixes_applied_count": len(fixes_applied) if fixes_applied else 0,
        },
        "issues_by_category": issues_by_category,
        "issues": issues_flat,
    }

    if fixes_applied:
        data["fixes_applied"] = [
            {
                "key": fix.get("key", ""),
                "label": fix.get("label", ""),
                "previous_translation": fix.get("previous_translation", ""),
                "fixed_translation": fix.get("fixed_translation", ""),
                "issue_category": fix.get("issue_category", ""),
                "fix_description": fix.get("fix_description", ""),
            }
            for fix in fixes_applied
        ]

    path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")


def export_html(
    report: ValidationReport,
    path: Path,
    fixes_applied: Optional[List[dict]] = None,
) -> None:
    """Write a standalone HTML report with embedded CSS.

    If *fixes_applied* is provided, an additional "Fixes Applied" table
    section is rendered below the issues table.
    """
    path = Path(path)
    error_count = len(report.errors)
    warning_count = len(report.warnings)

    # Group issues by category
    by_category: dict[str, list] = {}
    for issue in report.issues:
        by_category.setdefault(issue.category, []).append(issue)

    # Build table rows grouped by category
    table_rows = []
    for category, issues in by_category.items():
        table_rows.append(
            f'<tr class="category-header"><td colspan="5">'
            f'{html_escape(category)} ({len(issues)} issue(s))</td></tr>'
        )
        for issue in issues:
            component = issue.component if issue.component else _component_from_key(issue.key)
            sev_class = issue.severity
            table_rows.append(
                f"<tr>"
                f'<td class="{sev_class}">{html_escape(issue.severity)}</td>'
                f"<td>{html_escape(issue.category)}</td>"
                f"<td>{html_escape(component)}</td>"
                f"<td>{html_escape(issue.key)}</td>"
                f"<td>{html_escape(issue.message)}</td>"
                f"</tr>"
            )

    rows_html = "\n".join(table_rows)

    # Build fixes table if available
    fixes_section = ""
    if fixes_applied:
        fix_rows = []
        for fix in fixes_applied:
            fix_rows.append(
                f"<tr>"
                f"<td>{html_escape(fix.get('key', ''))}</td>"
                f"<td>{html_escape(fix.get('label', ''))}</td>"
                f"<td>{html_escape(fix.get('previous_translation', ''))}</td>"
                f"<td>{html_escape(fix.get('fixed_translation', ''))}</td>"
                f"<td>{html_escape(fix.get('issue_category', ''))}</td>"
                f"<td>{html_escape(fix.get('fix_description', ''))}</td>"
                f"</tr>"
            )
        fix_rows_html = "\n".join(fix_rows)
        fixes_section = f"""
<h2>Fixes Applied ({len(fixes_applied)})</h2>
<table>
<thead>
<tr><th>Key</th><th>Label</th><th>Previous Translation</th><th>Fixed Translation</th><th>Issue Category</th><th>Fix Applied</th></tr>
</thead>
<tbody>
{fix_rows_html}
</tbody>
</table>
"""

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>Validation Report</title>
<style>
body {{ font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif; margin: 2em; color: #1a1a1a; }}
h1 {{ color: #1e3a5f; }}
h2 {{ color: #1e3a5f; margin-top: 2em; }}
.summary {{ background: #f0f4f8; padding: 1em; border-radius: 6px; margin-bottom: 1.5em; }}
.summary span {{ margin-right: 2em; font-weight: 600; }}
.errors {{ color: #dc2626; }}
.warnings {{ color: #d97706; }}
table {{ border-collapse: collapse; width: 100%; margin-top: 1em; }}
th, td {{ border: 1px solid #d1d5db; padding: 0.5em 0.75em; text-align: left; }}
th {{ background: #1e3a5f; color: white; }}
tr:nth-child(even) {{ background: #f9fafb; }}
tr.category-header {{ background: #e2e8f0; font-weight: 700; }}
td.error {{ color: #dc2626; font-weight: 600; }}
td.warning {{ color: #d97706; font-weight: 600; }}
td.info {{ color: #2563eb; font-weight: 600; }}
</style>
</head>
<body>
<h1>Validation Report</h1>
<div class="summary">
<span class="errors">Errors: {error_count}</span>
<span class="warnings">Warnings: {warning_count}</span>
<span>Total: {len(report.issues)}</span>
</div>
<table>
<thead>
<tr><th>Severity</th><th>Category</th><th>Component</th><th>Key</th><th>Message</th></tr>
</thead>
<tbody>
{rows_html}
</tbody>
</table>
{fixes_section}
</body>
</html>
"""
    path.write_text(html, encoding="utf-8")


def export_xlsx(
    report: ValidationReport,
    path: Path,
    fixes_applied: Optional[List[dict]] = None,
    *,
    document_stats: Optional[dict] = None,
    document_name: Optional[str] = None,
    all_issues: Optional[List["ValidationIssue"]] = None,
) -> None:
    """Write a multi-sheet Excel report with Summary + per-category sheets.

    Sheet layout:
      - **Summary**: Document info header, category breakdown table with
        error/warning/total counts, and a TOTAL row.
      - **One sheet per category**: Each issue category gets its own sheet
        with columns: #, Severity, Key, Component, Source Label, Translation,
        Message.
      - **Fixes Applied** (if any): Details of auto-fix changes.

    Parameters
    ----------
    report : ValidationReport
        The core validation report (used when *all_issues* is not provided).
    path : Path
        Destination file path.
    fixes_applied : list of dict, optional
        Auto-fix details to include in a "Fixes Applied" sheet.
    document_stats : dict, optional
        Output of ``Document.stats()`` with keys: total, translated,
        untranslated.  Used for the Summary header section.
    document_name : str, optional
        Human-readable document/file name shown in the Summary header.
    all_issues : list of ValidationIssue, optional
        Full issues list (including translation_failed entries added by the
        GUI).  When not provided, falls back to ``report.issues``.
    """
    from datetime import datetime

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    bold_font = Font(bold=True)

    # Use the full issues list if provided (includes translation_failed etc.)
    issues = all_issues if all_issues is not None else list(report.issues)

    wb = Workbook()

    # --- Sheet 1: Summary ---
    ws_summary = wb.active
    ws_summary.title = "Summary"

    # Header info section
    ws_summary.append(["Validation Report"])
    ws_summary["A1"].font = Font(bold=True, size=14)
    ws_summary.append(["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    if document_name:
        ws_summary.append(["Document", document_name])
    if document_stats:
        ws_summary.append(["Total Rows", document_stats.get("total", "")])
        ws_summary.append(["Translated", document_stats.get("translated", "")])
        ws_summary.append(["Untranslated", document_stats.get("untranslated", "")])
    ws_summary.append([])  # blank separator row

    # Category breakdown table
    summary_header_row = ws_summary.max_row + 1
    ws_summary.append(["Category", "Errors", "Warnings", "Total", "Auto-fixable"])

    # Style the category table header
    for col_idx in range(1, 6):
        cell = ws_summary.cell(row=summary_header_row, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill

    # Group issues by category and count errors/warnings
    categories: dict[str, dict[str, int]] = {}
    for issue in issues:
        cat = issue.category
        if cat not in categories:
            categories[cat] = {"errors": 0, "warnings": 0}
        if issue.severity == "error":
            categories[cat]["errors"] += 1
        else:
            categories[cat]["warnings"] += 1

    # Determine auto-fixable categories (categories that the autofix module
    # can handle deterministically).
    auto_fixable_categories = {
        "length_limit", "duplicate_key", "token_drift",
        "html_mismatch", "missing_placeholder", "missing_message_format",
        "whitespace_only",
    }

    total_errors = 0
    total_warnings = 0
    total_autofixable = 0
    for cat in sorted(categories.keys()):
        counts = categories[cat]
        cat_total = counts["errors"] + counts["warnings"]
        autofixable = cat_total if cat in auto_fixable_categories else 0
        ws_summary.append([cat, counts["errors"], counts["warnings"], cat_total, autofixable])
        total_errors += counts["errors"]
        total_warnings += counts["warnings"]
        total_autofixable += autofixable

    # TOTAL row (bold)
    total_row_idx = ws_summary.max_row + 1
    ws_summary.append(["TOTAL", total_errors, total_warnings,
                       total_errors + total_warnings, total_autofixable])
    for col_idx in range(1, 6):
        ws_summary.cell(row=total_row_idx, column=col_idx).font = bold_font

    # Auto-size summary columns
    _xlsx_autosize(ws_summary, 5)

    # --- Per-category sheets ---
    for cat in sorted(categories.keys()):
        # Excel sheet names are limited to 31 characters
        sheet_title = cat[:31]
        ws = wb.create_sheet(title=sheet_title)
        cat_columns = ["#", "Severity", "Key", "Component",
                       "Source Label", "Translation", "Message"]
        ws.append(cat_columns)
        _xlsx_style_header(ws, len(cat_columns), header_font, header_fill)

        cat_issues = [i for i in issues if i.category == cat]
        for idx, issue in enumerate(cat_issues, 1):
            component = issue.component if issue.component else _component_from_key(issue.key)
            ws.append([
                idx,
                issue.severity,
                issue.key,
                component,
                "",  # Source Label - filled below if document available
                "",  # Translation - filled below if document available
                issue.message,
            ])

        _xlsx_autosize(ws, len(cat_columns))

    # --- Fixes Applied sheet ---
    if fixes_applied:
        ws_fixes = wb.create_sheet("Fixes Applied")
        fix_columns = [
            "Key", "Label", "Previous Translation",
            "Fixed Translation", "Issue Category", "Fix Applied",
        ]
        ws_fixes.append(fix_columns)
        _xlsx_style_header(ws_fixes, len(fix_columns), header_font, header_fill)

        for fix in fixes_applied:
            ws_fixes.append([
                fix.get("key", ""),
                fix.get("label", ""),
                fix.get("previous_translation", ""),
                fix.get("fixed_translation", ""),
                fix.get("issue_category", ""),
                fix.get("fix_description", ""),
            ])

        _xlsx_autosize(ws_fixes, len(fix_columns))

    wb.save(path)


def _xlsx_style_header(ws, col_count: int, font: "Font", fill: "PatternFill") -> None:
    """Apply header styling and freeze panes at A2."""
    for col_idx in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = font
        cell.fill = fill
    ws.freeze_panes = "A2"


def _xlsx_autosize(ws, col_count: int, max_width: int = 80) -> None:
    """Auto-size columns based on content, capped at *max_width*."""
    from openpyxl.utils import get_column_letter

    widths = [len(str(ws.cell(row=1, column=c).value or "")) for c in range(1, col_count + 1)]
    for row in ws.iter_rows(min_row=2, max_col=col_count, values_only=True):
        for idx, value in enumerate(row):
            if value is None:
                continue
            length = len(str(value))
            if length > widths[idx]:
                widths[idx] = length
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = min(width + 2, max_width)
