"""Parse an existing Excel file to extract source -> translation mappings.

This module supports importing translations from previously translated
workbooks so they can be reused in a new translation run with highest
priority (before TM, before network translation).

Supported layouts
-----------------
1. **Standard STF-Excel output** -- has Key, Label, Translation columns.
   The Label column is the source text and Translation is the target.
2. **Glossary-style** -- two columns: Source/Target (or similar column
   names like Source/Translation).
3. **Custom columns** -- caller specifies ``source_col`` and
   ``translation_col`` by exact header name.

Auto-detection tries column names in this order:
- "Label" + "Translation"
- "Source" + "Translation"
- "Source" + "Target"
- First two text columns if only 2 columns exist (glossary-style)
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, Optional

from openpyxl import load_workbook


@dataclass
class ImportedTranslations:
    """Holds a mapping of source text (label) to translation."""

    translations: Dict[str, str] = field(default_factory=dict)

    @property
    def count(self) -> int:
        return len(self.translations)

    def get(self, label: str) -> Optional[str]:
        """Look up a translation for the given label text."""
        return self.translations.get(label)


@dataclass
class ImportResult:
    """Result of parsing an import file."""

    total_imported: int = 0


def parse_translation_file(
    path: Path | str,
    *,
    source_col: Optional[str] = None,
    translation_col: Optional[str] = None,
) -> ImportedTranslations:
    """Load an Excel file and extract source -> translation mappings.

    Parameters
    ----------
    path:
        Path to an ``.xlsx`` file.
    source_col:
        Explicit column header name for the source text.  When provided
        together with ``translation_col``, auto-detection is skipped.
    translation_col:
        Explicit column header name for the translation text.

    Returns
    -------
    ImportedTranslations
        Object holding the extracted mappings.
    """
    path = Path(path)
    if not path.exists():
        return ImportedTranslations()

    wb = load_workbook(path, data_only=True, read_only=True)
    translations: Dict[str, str] = {}

    for ws in wb.worksheets:
        sheet_translations = _extract_from_sheet(
            ws, source_col=source_col, translation_col=translation_col
        )
        # Later sheets do not overwrite earlier sheets (first occurrence wins).
        for src, tgt in sheet_translations.items():
            if src not in translations:
                translations[src] = tgt

    wb.close()
    return ImportedTranslations(translations=translations)


def _extract_from_sheet(
    ws,
    *,
    source_col: Optional[str] = None,
    translation_col: Optional[str] = None,
) -> Dict[str, str]:
    """Extract source -> translation pairs from a single worksheet."""
    rows = ws.iter_rows(values_only=True)
    try:
        header_row = list(next(rows))
    except StopIteration:
        return {}

    headers = [_stringify(h) for h in header_row]

    src_idx, tgt_idx = _detect_columns(headers, source_col, translation_col)
    if src_idx is None or tgt_idx is None:
        return {}

    translations: Dict[str, str] = {}
    for row in rows:
        if row is None:
            continue
        cells = list(row)
        if src_idx >= len(cells) or tgt_idx >= len(cells):
            continue
        source = _stringify(cells[src_idx])
        translation = _stringify(cells[tgt_idx])
        if source and translation:
            # First occurrence wins within a sheet too.
            if source not in translations:
                translations[source] = translation

    return translations


def _detect_columns(
    headers: list[str],
    source_col: Optional[str],
    translation_col: Optional[str],
) -> tuple[Optional[int], Optional[int]]:
    """Detect source and translation column indices.

    Returns (source_index, translation_index) or (None, None) if not found.
    """
    # Custom columns take priority.
    if source_col and translation_col:
        src_idx = _find_header(headers, source_col)
        tgt_idx = _find_header(headers, translation_col)
        if src_idx is not None and tgt_idx is not None:
            return src_idx, tgt_idx

    # Strategy 1: Label + Translation (standard STF-Excel output)
    label_idx = _find_header(headers, "Label")
    trans_idx = _find_header(headers, "Translation")
    if label_idx is not None and trans_idx is not None:
        return label_idx, trans_idx

    # Strategy 2: Source + Translation
    source_idx = _find_header(headers, "Source")
    if source_idx is not None and trans_idx is not None:
        return source_idx, trans_idx

    # Strategy 3: Source + Target
    target_idx = _find_header(headers, "Target")
    if source_idx is not None and target_idx is not None:
        return source_idx, target_idx

    # Strategy 4: Glossary-style -- exactly 2 non-empty columns
    non_empty = [(i, h) for i, h in enumerate(headers) if h.strip()]
    if len(non_empty) == 2:
        return non_empty[0][0], non_empty[1][0]

    return None, None


def _find_header(headers: list[str], name: str) -> Optional[int]:
    """Case-insensitive header lookup."""
    name_lower = name.lower()
    for i, h in enumerate(headers):
        if h.strip().lower() == name_lower:
            return i
    return None


def _stringify(value) -> str:
    """Convert a cell value to a clean string."""
    if value is None:
        return ""
    text = str(value)
    if text.lower() in {"nan", "none"}:
        return ""
    return text
