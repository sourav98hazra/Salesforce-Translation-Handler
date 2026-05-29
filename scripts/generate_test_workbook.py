#!/usr/bin/env python3
"""Generate a manual-testing + app-guide Excel workbook.

Produces ``Manual_Test_Workbook.xlsx`` at the repo root with five sheets:

  1. Start Here        -- what the app is + how to launch it
  2. App Guide         -- the 6 phases, what each does, where things live
  3. Manual Test Cases -- a checklist of every scenario with Pass/Fail dropdowns
  4. Test Data         -- ready-to-paste tricky inputs (SFIDs, placeholders, HTML, ...)
  5. Feature Reference -- feature -> where to find it / how to use it

Re-generate any time with:

    python scripts/generate_test_workbook.py
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "Manual_Test_Workbook.xlsx"

# ---- shared styles --------------------------------------------------------
ACCENT = "1F4E78"
ACCENT_LIGHT = "DCE6F1"
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor=ACCENT)
TITLE_FONT = Font(bold=True, size=16, color=ACCENT)
SUB_FONT = Font(bold=True, size=12, color=ACCENT)
WRAP = Alignment(vertical="top", wrap_text=True)
TOP = Alignment(vertical="top")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
ZEBRA = PatternFill("solid", fgColor="F2F6FB")


def _style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP
        cell.border = BORDER
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def _widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_table(ws, headers, rows, widths, start_row=1, zebra=True):
    for i, h in enumerate(headers, start=1):
        ws.cell(row=start_row, column=i, value=h)
    _style_header(ws, len(headers), row=start_row)
    _widths(ws, widths)
    r = start_row + 1
    for ri, row in enumerate(rows):
        for ci, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=ci, value=val)
            cell.alignment = WRAP
            cell.border = BORDER
            if zebra and ri % 2 == 1:
                cell.fill = ZEBRA
        r += 1
    return r


# ---------------------------------------------------------------------------
# Sheet 1: Start Here
# ---------------------------------------------------------------------------
def sheet_start_here(wb):
    ws = wb.active
    ws.title = "Start Here"
    _widths(ws, [4, 60, 60])
    ws["B2"] = "Salesforce Translation Manager"
    ws["B2"].font = TITLE_FONT
    ws["B3"] = "Manual Test Workbook & App Guide"
    ws["B3"].font = SUB_FONT

    blocks = [
        ("What is this app?",
         "A desktop app (also CLI + Python library) that takes a Salesforce "
         "Translation Workbench .stf export, runs it through a 6-phase "
         "pipeline (Import -> Excel -> Translate -> Browse & Review -> "
         "Validate & Fix -> Export STF), and writes back Salesforce-ready "
         ".stf files. It protects Salesforce IDs, {!placeholders}, "
         "MessageFormat tokens, URLs, emails and HTML so the translator "
         "never corrupts them."),
        ("How do I launch it? (Windows, easiest)",
         "Double-click setup_desktop_app.bat once. It builds the app and puts "
         "a 'Salesforce Translation Manager' shortcut on your Desktop. After "
         "that, just double-click that Desktop shortcut to open the app."),
        ("How do I launch it? (any OS, developer)",
         "pip install -e \".[gui]\"  then run:  stx-app   (GUI)  or  "
         "stx --help  (command line)."),
        ("How do I use THIS workbook?",
         "Go to the 'Manual Test Cases' tab. Work top to bottom. For each "
         "row, follow Steps using the matching row from the 'Test Data' tab, "
         "compare what you see against 'Expected Result', then pick "
         "Pass/Fail/Blocked in the Result column. Put your name + date and "
         "any notes. The 'App Guide' and 'Feature Reference' tabs explain "
         "every screen so you always know where you are."),
        ("Where are my output files?",
         "Excel: wherever you chose to save (default <source>_organized.xlsx, "
         "_translated.xlsx, etc.). STF: the output folder you pick in Phase 6 "
         "(Super_STF_<code>.stf, TranslatedOnly_STF_<code>.stf, "
         "UntranslatedOnly_STF_<code>.stf)."),
        ("Tip",
         "Every phase works on its own — you can jump straight to Phase 5 to "
         "validate someone else's workbook, or Phase 6 to convert a "
         "translated Excel directly to STF, without doing the earlier steps."),
    ]
    r = 5
    for title, body in blocks:
        ws.cell(row=r, column=2, value=title).font = SUB_FONT
        b = ws.cell(row=r, column=3, value=body)
        b.alignment = WRAP
        ws.row_dimensions[r].height = 78
        r += 1


# ---------------------------------------------------------------------------
# Sheet 2: App Guide
# ---------------------------------------------------------------------------
def sheet_app_guide(wb):
    ws = wb.create_sheet("App Guide")
    headers = ["Phase", "Screen name", "What it does", "Key buttons / controls",
               "Output / result"]
    rows = [
        ["1", "Import STF",
         "Parses the Salesforce .stf you exported. Shows language, code, type, "
         "row counts.",
         "Browse STF...; Save copy as STF...; Continue to Phase 2",
         "In-memory document; optional clean STF copy"],
        ["2", "STF -> Excel",
         "Groups rows by component type into an organised workbook (one sheet "
         "per component) plus a Content Details index.",
         "Convert and save .xlsx; Save copy to...; Load existing .xlsx",
         "<source>_organized.xlsx"],
        ["3", "Translate",
         "Auto-translates untranslated rows. Protects IDs/placeholders/HTML. "
         "Live EN->target feed with counters + ETA.",
         "Source/Target language; Filter Components...; Start translation; "
         "Save copy to...",
         "<source>_translated.xlsx with audit sheets"],
        ["4", "Browse & Review",
         "Browse + edit translations. Auto-validates on entry. Side-by-side "
         "Source/Translation editor (drag the divider to resize).",
         "Search/filter; Apply; Reset to source; Load reviewed Excel; Save",
         "<source>_reviewed.xlsx"],
        ["5", "Validate & Fix",
         "Shows ONLY rows with issues. Deterministic auto-fixers + manual edit.",
         "Auto-fix all; Re-validate; inline editor; double-click row -> Phase 4",
         "<source>_fixed.xlsx"],
        ["6", "Export STF",
         "Writes the 3 Salesforce-ready .stf files. Also a direct 'load "
         "translated Excel -> STF' entry point.",
         "Language/Code; Output dir; Run validation (advisory); Export 3 STF",
         "Super_/TranslatedOnly_/UntranslatedOnly_STF_<code>.stf"],
        ["—", "Settings (Edit > Settings, Ctrl+,)",
         "All advanced options: backend + API key, workers, rate limit, "
         "wake-lock, glossary path, TM path, batch targets, theme.",
         "Translation / Resources / Appearance tabs",
         "Saved across sessions"],
        ["—", "Help menu",
         "In-app User Guide (F1) and About dialog.",
         "User Guide; About",
         "—"],
    ]
    _write_table(ws, headers, rows, [8, 26, 46, 40, 34])
    for row in range(2, 2 + len(rows)):
        ws.row_dimensions[row].height = 70


# ---------------------------------------------------------------------------
# Sheet 3: Manual Test Cases
# ---------------------------------------------------------------------------
def sheet_test_cases(wb):
    ws = wb.create_sheet("Manual Test Cases")
    headers = ["ID", "Area", "Scenario", "Steps to perform", "Test data (tab)",
               "Expected result", "Result", "Tester", "Date", "Notes"]

    rows = [
        # Launch / setup
        ["T-01", "Launch", "App launches from Desktop shortcut",
         "Double-click the Desktop shortcut created by setup_desktop_app.bat",
         "—", "App window opens within a few seconds; no error dialog", "", "", "", ""],
        ["T-02", "Launch", "App launches via command",
         "Run 'stx-app' in a terminal from the project folder",
         "—", "Main window opens on Phase 1", "", "", "", ""],
        ["T-03", "Window", "Window fits the screen",
         "Open the app on a small / laptop screen",
         "—", "Window opens fully on-screen, not wider than the display", "", "", "", ""],
        ["T-04", "Window", "Window can be resized narrow",
         "Drag the right/bottom edge to shrink the window",
         "—", "Window shrinks to ~900x600 min; sidebar stays readable", "", "", "", ""],
        ["T-05", "Theme", "All 5 themes apply",
         "Edit > Settings > Appearance; try Light/Dark/Ocean/Forest/Sunset/Auto",
         "—", "Colours change immediately; text stays readable", "", "", "", ""],
        # Phase 1
        ["T-06", "Phase 1", "Import a valid STF",
         "Phase 1 > Browse STF... > pick the sample .stf",
         "STF Samples", "Language, code, type and row counts shown correctly", "", "", "", ""],
        ["T-07", "Phase 1", "Metadata grid + tooltips",
         "Hover over each metadata field",
         "STF Samples", "Full value shown in tooltip if truncated", "", "", "", ""],
        ["T-08", "Phase 1", "Save clean STF copy",
         "Click 'Save copy as STF...'; choose a path",
         "STF Samples", "A .stf file is written and re-opens correctly", "", "", "", ""],
        ["T-09", "Phase 1", "Drag-and-drop an .stf",
         "Drag a .stf file anywhere onto the window",
         "STF Samples", "App routes to Phase 1 and parses it", "", "", "", ""],
        # Phase 2
        ["T-10", "Phase 2", "Convert to organised Excel",
         "Phase 2 > Convert and save .xlsx",
         "—", "Workbook created: one sheet per component + Content Details", "", "", "", ""],
        ["T-11", "Phase 2", "Salesforce IDs stay text",
         "Open the saved .xlsx; check an ID like 001D000000IqhSL and 007",
         "Tricky Labels", "IDs are NOT reformatted to numbers/scientific", "", "", "", ""],
        ["T-12", "Phase 2", "Formula injection guarded",
         "Include a label starting with = + - @ ; export; open in Excel",
         "Tricky Labels", "Cell shows the literal text, no formula executes", "", "", "", ""],
        ["T-13", "Phase 2", "Load existing organised .xlsx",
         "Phase 2 > Load existing organised .xlsx",
         "—", "Document loads; you can continue from here", "", "", "", ""],
        # Phase 3
        ["T-14", "Phase 3", "Pick source/target language",
         "Set Source=English, Target=Japanese",
         "—", "Languages accepted; estimate updates", "", "", "", ""],
        ["T-15", "Phase 3", "Filter components dialog",
         "Filter Components... > search, status filter, Select all/none/Invert",
         "—", "Selection count + estimated rows update live", "", "", "", ""],
        ["T-16", "Phase 3", "Translate runs with live feed",
         "Start translation; watch the feed",
         "—", "Each line shows EN->JA pair + counters; ETA shown; ends with DONE summary", "", "", "", ""],
        ["T-17", "Phase 3", "Placeholders/IDs/HTML preserved",
         "After translating, inspect rows with {!..}, IDs, <tags>",
         "Tricky Labels", "All tokens identical to source; only prose translated", "", "", "", ""],
        ["T-18", "Phase 3", "Cannot start twice / cancel safely",
         "Click Start rapidly; then Cancel",
         "—", "Only one run starts; cancel stops cleanly, no crash", "", "", "", ""],
        ["T-19", "Phase 3", "Translation memory speeds re-run",
         "Set a TM path in Settings; translate; clear results; translate again",
         "—", "Second run is near-instant; feed shows TM hits", "", "", "", ""],
        ["T-20", "Phase 3", "Glossary do-not-translate",
         "Add a glossary with Bayer=DNT; translate text containing 'Bayer'",
         "Glossary CSV", "'Bayer' is unchanged in the translation", "", "", "", ""],
        ["T-21", "Phase 3", "Multi-language batch",
         "Settings > batch targets 'fr,de,es'; run",
         "—", "Output produced for each target language", "", "", "", ""],
        # Phase 4
        ["T-22", "Phase 4", "Browse + filter",
         "Use component/status dropdowns + search box",
         "—", "Table filters immediately and correctly", "", "", "", ""],
        ["T-23", "Phase 4", "Edit a translation",
         "Select a row; edit on the right; click Apply",
         "—", "Change saved to the document; counters update", "", "", "", ""],
        ["T-24", "Phase 4", "Resize editor (drag divider)",
         "Drag the horizontal divider between the table and the editor",
         "—", "Editor grows; Source/Translation boxes get taller", "", "", "", ""],
        ["T-25", "Phase 4", "Reset to source",
         "Click 'Reset to source' on an edited row",
         "—", "Translation reverts to the source label", "", "", "", ""],
        ["T-26", "Phase 4", "Re-upload edited Excel",
         "Edit the .xlsx in Excel, save, then Load reviewed Excel",
         "—", "In-memory document replaced with your edits", "", "", "", ""],
        # Phase 5
        ["T-27", "Phase 5", "Only issue rows shown",
         "Enter Phase 5 with a doc that has issues",
         "Tricky Labels", "Clean rows hidden; only problem rows listed with severity", "", "", "", ""],
        ["T-28", "Phase 5", "Duplicate keys flagged",
         "Include two rows with the same key; validate",
         "Tricky Labels", "duplicate_key error shown", "", "", "", ""],
        ["T-29", "Phase 5", "Length overflow flagged + trimmed",
         "CustomField translation > 80 chars; Auto-fix all",
         "Tricky Labels", "length_limit error; auto-fix trims at word boundary + …", "", "", "", ""],
        ["T-30", "Phase 5", "Lost placeholder restored",
         "Translation missing a {!..}; Auto-fix all",
         "Tricky Labels", "token_drift flagged; placeholder restored", "", "", "", ""],
        ["T-31", "Phase 5", "Re-validate clears issues",
         "After Auto-fix all, click Re-validate",
         "—", "Issue count drops; clean rows confirmed", "", "", "", ""],
        ["T-32", "Phase 5", "Jump to Phase 4 for context",
         "Double-click an issue row",
         "—", "Lands on that row in Phase 4 with filters cleared", "", "", "", ""],
        # Phase 6
        ["T-33", "Phase 6", "Export 3 STF files",
         "Set language/code + output dir; Export 3 STF files",
         "—", "Super_/TranslatedOnly_/UntranslatedOnly_STF_<code>.stf created", "", "", "", ""],
        ["T-34", "Phase 6", "STF format is correct",
         "Open Super_STF in a text editor; check separators + encoding",
         "—", "UTF-8, LF endings, no BOM; exact TRANSLATED/OUTDATED separators", "", "", "", ""],
        ["T-35", "Phase 6", "Direct convert (skip phases)",
         "Open app > Phase 6 > Load translated Excel... > Export",
         "—", "STF produced without doing Phases 1-5", "", "", "", ""],
        ["T-36", "Phase 6", "Round-trip integrity",
         "Re-import the exported Super_STF into Phase 1",
         "—", "All keys/labels/translations match the original", "", "", "", ""],
        # Cross-cutting
        ["T-37", "General", "Status log visible + toggle",
         "View > Show Status Log (Ctrl+L); watch messages during actions",
         "—", "Log panel shows/hides; heading 'Status log' readable", "", "", "", ""],
        ["T-38", "General", "Recent files menu",
         "File > Recent files after opening a few files",
         "—", "Recently used files listed; clicking re-opens", "", "", "", ""],
        ["T-39", "General", "Keyboard shortcuts",
         "Try Ctrl+0..5, Ctrl+O, Ctrl+S, Ctrl+, , F1, Ctrl+Q",
         "—", "Each shortcut performs its action", "", "", "", ""],
        ["T-40", "General", "Pop-out panels",
         "Use the pop-out (↗) on Preview (Phase 1) / Live feed (Phase 3)",
         "—", "Panel detaches into its own window", "", "", "", ""],
        ["T-41", "Help", "User Guide + About open",
         "Help > User Guide (F1); Help > About",
         "—", "Both dialogs open and render; version shown in About", "", "", "", ""],
        ["T-42", "Error", "Bad/corrupt file handled",
         "Try to import a non-STF / empty file",
         "STF Samples", "Friendly message, no crash", "", "", "", ""],
        # CLI
        ["T-43", "CLI", "stx info",
         "Run: stx info <sample.stf>",
         "STF Samples", "Prints language + counts, exit code 0", "", "", "", ""],
        ["T-44", "CLI", "stx run end-to-end (no network)",
         "Run: stx run <sample.stf> ./out --skip-translation",
         "STF Samples", "Excel + STF artifacts produced in ./out", "", "", "", ""],
        ["T-45", "CLI", "stx validate",
         "Run: stx validate <sample.stf>",
         "STF Samples", "Issue report printed; exit reflects errors", "", "", "", ""],
        ["T-46", "CLI", "stx backends",
         "Run: stx backends",
         "—", "Lists google/deepl/azure/openai + key requirements", "", "", "", ""],
    ]
    last = _write_table(ws, headers, rows,
                        [7, 12, 30, 46, 14, 46, 11, 12, 12, 30])
    for row in range(2, last):
        ws.row_dimensions[row].height = 46

    # Result dropdown (Pass/Fail/Blocked/Not Run) on column G
    dv = DataValidation(type="list",
                        formula1='"Pass,Fail,Blocked,Not Run"',
                        allow_blank=True)
    dv.prompt = "Select the outcome"
    dv.promptTitle = "Result"
    ws.add_data_validation(dv)
    dv.add(f"G2:G{last - 1}")

    # Light fill on the editable Result column to invite input
    for r in range(2, last):
        ws.cell(row=r, column=7).fill = PatternFill("solid", fgColor="FFF7E6")

    ws.auto_filter.ref = f"A1:J{last - 1}"


# ---------------------------------------------------------------------------
# Sheet 4: Test Data (ready to paste)
# ---------------------------------------------------------------------------
def sheet_test_data(wb):
    ws = wb.create_sheet("Test Data")
    ws["A1"] = ("Ready-to-paste test inputs. Copy a Label into the app (or into "
                "an .stf / .xlsx) to exercise the matching test case.")
    ws["A1"].font = SUB_FONT
    ws.merge_cells("A1:D1")
    ws["A1"].alignment = WRAP
    ws.row_dimensions[1].height = 30

    headers = ["Key", "Label (source)", "What it tests", "Expected handling"]
    rows = [
        ["CustomLabel.PlainText", "Click here to continue",
         "Normal prose", "Fully translated"],
        ["CustomLabel.Placeholder", "Welcome back, {!User.FirstName}!",
         "{!..} placeholder", "Placeholder identical; prose translated"],
        ["CustomLabel.MsgFormat", "You have {0} new messages out of {1}",
         "MessageFormat {0}/{1}", "Tokens identical; order preserved"],
        ["CustomLabel.SFID15", "Record 001D000000IqhSL was updated",
         "15-char Salesforce ID", "ID unchanged"],
        ["CustomLabel.SFID18", "See 01JD00000080e3FMAQ for details",
         "18-char Salesforce ID", "ID unchanged"],
        ["CustomLabel.URL", "Visit https://help.salesforce.com/articleView?id=foo",
         "URL", "URL unchanged"],
        ["CustomLabel.Email", "Contact support@example.com for help",
         "Email", "Email unchanged"],
        ["CustomLabel.Acronym", "Open the API and check the URL and WO",
         "ALL-CAPS acronyms", "API/URL/WO unchanged"],
        ["CustomLabel.Escapes", "Line one\\nLine two\\tTabbed",
         "Escape sequences", "\\n and \\t preserved literally"],
        ["CustomLabel.Html", "<p>Please contact <b>Support</b> now</p>",
         "HTML tags", "Tags/attributes intact; only text translated"],
        ["CustomLabel.FormulaPlus", "+1 (800) 555-0100",
         "Leading + (formula injection)", "Stored as literal text in Excel"],
        ["CustomLabel.FormulaEq", "=SUM(A1:A2)",
         "Leading = (formula injection)", "Stored as literal text in Excel"],
        ["CustomField.Account.Name.FieldLabel", "x" * 120,
         "CustomField length > 80", "Flagged length_limit; auto-fix trims"],
        ["CustomLabel.DupKey", "First definition",
         "Duplicate key (pair with next row)", "duplicate_key error"],
        ["CustomLabel.DupKey", "Second definition",
         "Duplicate key (same key as above)", "duplicate_key error; dedup keeps last"],
        ["CustomLabel.Brand", "Welcome to Bayer support",
         "Glossary do-not-translate", "'Bayer' unchanged when DNT glossary loaded"],
        ["CustomLabel.Cjk", "在庫管理システム",
         "Non-ASCII / CJK source", "Preserved end-to-end"],
        ["CustomLabel.LongRich",
         "<p>" + ("Lorem ipsum dolor sit amet. " * 200) + "</p>",
         "Long rich text > 5000 chars", "Chunked + recombined, tags intact"],
    ]
    last = _write_table(ws, headers, rows, [40, 60, 30, 36], start_row=3)
    for row in range(4, last):
        ws.row_dimensions[row].height = 40

    # A small ready-made STF block they can paste into a .stf file
    r = last + 2
    ws.cell(row=r, column=1, value="Sample STF file content (paste into a .stf):").font = SUB_FONT
    r += 1
    stf_lines = [
        "# Language: Japanese",
        "Language code: ja",
        "Type: Outdated and untranslated",
        "Translation type: Metadata",
        "",
        "# KEY\tLABEL",
        "CustomApp.Inside_Sales\tInside Sales",
        "CustomApp.Sales_Leader\tSales Leader\tセールスリーダー\t-",
        "CustomLabel.Greeting\tWelcome back, {!User.FirstName}!",
        "CustomField.Account.Name.FieldLabel\tAccount Name",
    ]
    for line in stf_lines:
        ws.cell(row=r, column=1, value=line)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        r += 1


# ---------------------------------------------------------------------------
# Sheet 5: Glossary CSV sample + Feature Reference
# ---------------------------------------------------------------------------
def sheet_feature_reference(wb):
    ws = wb.create_sheet("Feature Reference")
    headers = ["Feature", "Where / how", "Notes"]
    rows = [
        ["Translator backends", "Settings > Translation > Backend",
         "Google (free, no key), DeepL, Azure, OpenAI (paid need API key)"],
        ["Translation Memory (TM)", "Settings > Resources > Translation memory",
         "SQLite cache; reuses past translations; big speed-up on re-runs"],
        ["Glossary", "Settings > Resources > Glossary",
         "CSV: source,target,do_not_translate. DNT protects brand names"],
        ["Component scope", "Phase 3 > Filter Components...",
         "Choose which component types/keys to translate; saved to .stxscope.json"],
        ["Multi-language batch", "Settings > Translation > batch targets",
         "Comma-separated codes, e.g. fr,de,es"],
        ["Workers / rate limit", "Settings > Translation",
         "Parallelism + req/sec; rate limit auto-tunes for free Google tier"],
        ["Wake-lock", "Settings > Translation > Prevent system sleep",
         "Stops idle sleep during long runs (lid-close still suspends)"],
        ["Auto-fix", "Phase 5 > Auto-fix all / this row",
         "Restores placeholders, trims length, dedups keys, fixes HTML pairs"],
        ["Validation", "Phase 4 (on entry) + Phase 5",
         "duplicate_key, length_limit, token_drift, html_mismatch"],
        ["Themes", "Edit > Settings > Appearance (or View menu)",
         "Light, Dark, Ocean, Forest, Sunset, Auto"],
        ["Pop-out panels", "↗ icon on Preview / Live feed",
         "Detach a panel into its own window"],
        ["Status log", "View > Show Status Log (Ctrl+L)",
         "Running log of actions; toggle to hide"],
        ["CLI", "Terminal: stx --help",
         "info, stf2xlsx, translate, xlsx2stf, validate, run, scope, backends"],
        ["Glossary CSV format", "(example below)",
         "Header: source,target,do_not_translate"],
    ]
    last = _write_table(ws, headers, rows, [26, 42, 56])
    for row in range(2, last):
        ws.row_dimensions[row].height = 32

    r = last + 2
    ws.cell(row=r, column=1, value="Sample glossary.csv:").font = SUB_FONT
    r += 1
    for line in ["source,target,do_not_translate",
                 "Bayer,,true",
                 "ATLS,,true",
                 "case,ケース,",
                 "record,レコード,"]:
        ws.cell(row=r, column=1, value=line)
        r += 1


def main() -> int:
    wb = Workbook()
    sheet_start_here(wb)
    sheet_app_guide(wb)
    sheet_test_cases(wb)
    sheet_test_data(wb)
    sheet_feature_reference(wb)
    wb.save(OUT)
    print(f"Wrote {OUT}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
